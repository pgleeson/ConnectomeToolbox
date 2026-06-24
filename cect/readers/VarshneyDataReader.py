from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections
from cect.ConnectomeDataset import ConnectomeDataset
from cect.ConnectomeDataset import get_dataset_source_on_github
from cect.ConnectomeDataset import LOAD_READERS_FROM_CACHE_BY_DEFAULT

from cect.Cells import UNSPECIFIED_BODY_WALL_MUSCLE

from cect.Neurotransmitters import GENERIC_CHEM_SYN_CLASS, CHEMICAL_SYN_TYPE
from cect.Neurotransmitters import GENERIC_ELEC_SYN_CLASS, ELECTRICAL_SYN_TYPE

from openpyxl import load_workbook

import os

from cect import print_

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/../data/"

spreadsheet_name = "NeuronConnectFormatted.xlsx"
spreadsheet_name = "NeuronConnect.xls"  # has old name...
filename = "%s%s" % (spreadsheet_location, spreadsheet_name)

NAME = "Varshney"

READER_DESCRIPTION = (
    """Data extracted from %s for neuronal connectivity"""
    % get_dataset_source_on_github(filename.split("/")[-1])
)

DATASET_DESCRIPTION = "A corrected and extended version of the White et al. chemical and electrical wiring diagram, incorporating original Mind of a Worm data, Durbin's unpublished reconstructions, new EM imaging of previously unimaged dorsal cord regions, and over 3,000 synapse additions or corrections, particularly in the ventral cord motor neuron connectivity."

WEIGHTS = "Weights are the total number of synaptic contacts from neuron A to neuron B. Contacts are given equal weight regardless of the apparent size of the synaptic apposition."

SEND_SYN = "S"
SEND_POLY_SYN = "Sp"
SEND_ANY = "SEND"
RECEIVE_SYN = "R"
RECEIVE_POLY_SYN = "Rp"
RECEIVE_ANY = "RECEIVE"

ELECT_JUNC_SYN = "EJ"

NMJ_ENDPOINT = "NMJ"


def _remove_leading_index_zero(cell):
    """
    Returns neuron name with an index without leading zero. E.g. VB01 -> VB1.
    """
    if cell[:2] in ["VA", "VB", "VC", "VD", "DA", "DB", "DD", "AS"] and cell[
        -2:
    ].startswith("0"):
        return "%s%s" % (cell[:-2], cell[-1:])
    return cell


class VarshneyDataReader(ConnectomeDataset):
    """Reader for Varshney et al. 2011 connectivity dataset"""

    def __init__(self):
        ConnectomeDataset.__init__(self)

        self.typed_conns = {
            SEND_SYN: [],
            SEND_POLY_SYN: [],
            RECEIVE_SYN: [],
            RECEIVE_POLY_SYN: [],
            SEND_ANY: set(),
            RECEIVE_ANY: set(),
            ELECT_JUNC_SYN: [],
            NMJ_ENDPOINT: [],
        }

        cells, neuron_conns = self.read_data()

        for conn in neuron_conns:
            self.add_connection_info(
                conn,
                append_existing_connections=True,
                check_overwritten_connections=True,
                fail_on_any_repeated_connection=False,
            )

    def _check_valid_synapse_type(self, syn_type):
        if syn_type not in self.typed_conns:
            raise ValueError(
                f"Synapse type '{syn_type}' not recognized for {NAME}. Valid types are: {list(self.typed_conns.keys())}"
            )
        return syn_type

    def read_data(self):
        cells = []
        conns = []

        if filename.endswith(".xls"):
            from xlrd import open_workbook

            wb = open_workbook(filename)
            rows = []
            sheet = wb.sheet_by_index(0)
            for row in range(1, sheet.nrows):
                rows.append(
                    (
                        str(sheet.cell(row, 0).value),
                        str(sheet.cell(row, 1).value),
                        str(sheet.cell(row, 2).value),
                        int(sheet.cell(row, 3).value),
                    )
                )
        else:
            wb = load_workbook(filename)
            sheet = wb.worksheets[0]
            rows = sheet.iter_rows(min_row=2, values_only=True)

        print_("Opened the Excel file: " + filename)

        for row in rows:  # Assuming data starts from the second row
            pre = str(row[0])
            post = str(row[1])

            pre = _remove_leading_index_zero(pre)
            post = _remove_leading_index_zero(post)

            if post == NMJ_ENDPOINT:
                post = UNSPECIFIED_BODY_WALL_MUSCLE

            syntype_here = self._check_valid_synapse_type(str(row[2]))
            num = int(row[3])

            self.typed_conns[syntype_here].append(f"{pre}_{post}_{num}")
            if syntype_here in [SEND_SYN, SEND_POLY_SYN]:
                self.typed_conns[SEND_ANY].add(f"{pre}_{post}")
            elif syntype_here in [RECEIVE_SYN, RECEIVE_POLY_SYN]:
                self.typed_conns[RECEIVE_ANY].add(f"{pre}_{post}")

            if not syntype_here == NMJ_ENDPOINT + "":
                synclass = (
                    GENERIC_ELEC_SYN_CLASS
                    if syntype_here == ELECT_JUNC_SYN
                    else GENERIC_CHEM_SYN_CLASS
                    if (syntype_here == SEND_POLY_SYN or syntype_here == SEND_SYN)
                    else GENERIC_CHEM_SYN_CLASS
                    if (syntype_here == NMJ_ENDPOINT)
                    else None
                )

                if syntype_here == ELECT_JUNC_SYN:
                    syntype = ELECTRICAL_SYN_TYPE
                else:
                    syntype = CHEMICAL_SYN_TYPE

                if synclass is not None:
                    conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                    if pre not in cells:
                        cells.append(pre)
                    if post not in cells:
                        cells.append(post)
                else:
                    if not (
                        syntype_here == RECEIVE_SYN or syntype_here == RECEIVE_POLY_SYN
                    ):
                        raise ValueError(
                            f"Warning: Unrecognized synapse type '{syntype_here}' for connection {pre} -> {post} for {NAME}."
                        )

        total = 0
        for syn_type, conn_list in self.typed_conns.items():
            total += len(conn_list)
            info = ""
            if syn_type in [SEND_SYN, SEND_POLY_SYN, RECEIVE_SYN, RECEIVE_POLY_SYN]:
                info = f"({', '.join(conn_list[:5])}..., {conn_list[-1]})"
            print_(f"  {syn_type}: {len(conn_list)} connections {info}")
        print_(f"  Total: {total} connections (half: {total / 2})")

        s_tot = len(self.typed_conns[SEND_SYN]) + len(self.typed_conns[SEND_POLY_SYN])

        r_tot = len(self.typed_conns[RECEIVE_SYN]) + len(
            self.typed_conns[RECEIVE_POLY_SYN]
        )

        print_(
            f"  Total chemical synapses: {s_tot} (send) + {r_tot} (receive) = {s_tot + r_tot}"
        )
        print_(f"  Total electrical synapses: {len(self.typed_conns[ELECT_JUNC_SYN])}")

        return cells, conns

    def read_muscle_data(self):
        conns = []
        neurons = []
        muscles = []

        return neurons, muscles, conns


def get_instance(from_cache=LOAD_READERS_FROM_CACHE_BY_DEFAULT):
    if from_cache:
        from cect.ConnectomeDataset import (
            load_connectome_dataset_file,
            get_cache_filename,
        )

        return load_connectome_dataset_file(get_cache_filename(__name__.split(".")[-1]))
    else:
        return VarshneyDataReader()


"""
read_data = my_instance.read_data
read_muscle_data = my_instance.read_muscle_data"""


def main():
    my_instance = get_instance()

    cells, neuron_conns = my_instance.read_data()
    neurons2muscles, muscles, muscle_conns = my_instance.read_muscle_data()

    analyse_connections(
        cells, neuron_conns, neurons2muscles, muscles, muscle_conns, print_details_on=[]
    )

    print_(" -- Finished analysing connections using: %s" % os.path.basename(__file__))

    cell = "ADAR"
    syntype = "Generic_CS"
    conns = my_instance.get_connections_from(cell, syntype)
    print(f"There are {len(conns)} connections from {cell} of type {syntype}:")
    for c in sorted(conns.keys()):
        print(f" {cell} -> {c}: {conns[c]}")

    import numpy as np

    for syntype in my_instance.connections.keys():
        conn_array = my_instance.connections[syntype]
        print(
            f" === Connection array for synapse type {syntype} has shape: {conn_array.shape}"
        )
        nonzero = np.count_nonzero(conn_array)
        count_diagonal_entries = np.count_nonzero(np.diag(conn_array))
        is_symmetric = np.array_equal(conn_array, conn_array.T)
        sym_info = ""
        if is_symmetric:
            diagonal_sum = np.sum(np.diag(conn_array))
            unique_sum = (np.sum(conn_array) - diagonal_sum) / 2 + diagonal_sum

            sym_info = f"Matrix is symmetric with <b>{count_diagonal_entries + int((nonzero - count_diagonal_entries) / 2)}</b> unique pairs (total unique pair weight: <b>{unique_sum}</b>)<br/>"

        diag_info = (
            "<b>%s</b> nodes with self-connections<br/>" % count_diagonal_entries
            if count_diagonal_entries > 0
            else ""
        )
        print(
            f"<b>{nonzero}</b> non-zero entries<br/>{diag_info}{sym_info}Avg. weight: <b>{np.mean(conn_array[conn_array != 0])}</b><br/>Sum of weights: <b>{np.sum(conn_array)}</b>"
        )


if __name__ == "__main__":
    main()
